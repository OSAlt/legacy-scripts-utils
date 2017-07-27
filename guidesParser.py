import string

import openpyxl

from script_lib.base_parser import BaseParser
from models.guide_model  import GuideModel


class Parser(BaseParser):
    def __init__(self):
        BaseParser.__init__(self, 'config/config.yaml')


    def read_sheet(self):
        wb = openpyxl.load_workbook('responses/guideResponses.xlsx')
        sheet = wb.get_sheet_by_name('Form Responses 1')
        data = []
        columns = list(string.ascii_uppercase)[:sheet.max_column]

        headers = [sheet['{}1'.format(x)].value for x in columns]

        # This is the stupid lib api, there has to be a better way of doing this.
        for i in range(1, sheet.max_row):
            rows = list(sheet.rows)[i]
            payload = {}
            for y in range(0, len(headers)):
                payload[headers[y]] = rows[y].value
            data.append(GuideModel(payload))

        self.sheet_data = data

    def run(self):
        (engine, session) = self.get_alchemy_connection()
        self.read_sheet()

        # u = self.sheet_data[0].persist(engine)
        user_entries = [x.persist(engine) for x in self.sheet_data]


def main():
    parser = Parser()
    parser.run()


if __name__ == '__main__':
    main()
