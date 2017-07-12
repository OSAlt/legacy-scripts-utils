import openpyxl
import string
from user import User


def read_sheet():
    wb = openpyxl.load_workbook('responses/responses.xlsx')
    sheet = wb.get_sheet_by_name('Form Responses 1')
    data = []
    columns = list(string.ascii_uppercase)[:sheet.max_column]

    headers = [sheet['{}1'.format(x)].value for x in columns ]


    ## This is the stupid lib api, there has to be a better way of doing this.
    for i in range(1, sheet.max_row):
        rows = list(sheet.rows)[i]
        payload = {}
        for y in range(0, len(headers)):
            payload[headers[y]] = rows[y].value
        data.append(payload)
    return data


def main():
    data = read_sheet()
    u = User(data[0])
    user_entries = [User(*x) for x in data]





if __name__ == '__main__':
    main()