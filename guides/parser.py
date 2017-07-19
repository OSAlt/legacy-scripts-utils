import string

import openpyxl
import yaml
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from models.user import User


class Parser(object):
    def __init__(self):
        self.db_config = None
        f = open('../config/config.yaml')
        data = yaml.load(f)
        f.close()
        self.config = data
        self.debug = self.config.get('debug', False)


    def __construct_dsn__(self):
        """
        generates a dsn connection string based on the yaml configuration.
        :return: string representation
        """
        db_settings = self.config['database']
        self.dsn = 'dbname={database} host={host} port={port} user={user} password={password}' \
            .format(**db_settings).replace('$HOSTNAME', db_settings['host'])

        self.db_config = {
            'drivername': db_settings.get('engine', 'mysql+pymysql'),
            'host': db_settings['host'],
            'port': db_settings['port'],
            'username': db_settings['user'],
            'password': db_settings['password'],
            'database': db_settings['database'],
            'schema':  db_settings['schema']

        }
        self.connection_string = "{engine}://{user}:{password}@{host}:{port}/{database}".format(**db_settings)

    def get_alchemy_connection(self):
        """
         Uses the DB settings to create a sqlalchemy engine + session.
        :return:
        """
        if self.db_config is None:
            self.__construct_dsn__()

        echo_sql = True if self.debug else False
        # engine = create_engine(URL(**self.db_config), echo=echo_sql, encoding=self.db_config.get('encoding', 'utf-8'))
        engine = create_engine(self.connection_string,  echo=echo_sql, encoding=self.db_config.get('encoding', 'utf8'))
        session = sessionmaker(bind=engine)
        return (engine, session)


    def read_sheet(self):
        wb = openpyxl.load_workbook('responses/responses.xlsx')
        sheet = wb.get_sheet_by_name('Form Responses 1')
        data = []
        columns = list(string.ascii_uppercase)[:sheet.max_column]

        headers = [sheet['{}1'.format(x)].value for x in columns]

        ## This is the stupid lib api, there has to be a better way of doing this.
        for i in range(1, sheet.max_row):
            rows = list(sheet.rows)[i]
            payload = {}
            for y in range(0, len(headers)):
                payload[headers[y]] = rows[y].value
            data.append(User(payload))

        self.sheet_data = data

    def run(self):
        (engine, session) = self.get_alchemy_connection()
        self.read_sheet()

        # u = self.sheet_data[0].persist(engine)
        user_entries = [x.persist(engine) for x in self.sheet_data]

def main():
    parser = Parser()
    parser.run()



if __name__ == '__main__':
    main()